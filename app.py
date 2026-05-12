"""
FastAPI backend for the GSRTC LED Screen proposal generator.
Serves the frontend UI and provides endpoints to generate PPT + Excel files.
"""

from fastapi import FastAPI, Request, Form, HTTPException
from fastapi.responses import HTMLResponse, Response, FileResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from typing import List
from pathlib import Path
import re

from generator import (
    generate_pptx, generate_xlsx, sanitize_filename,
    LOCATIONS_IN_ORDER, LOCATION_EXCEL_ROW
)
from openpyxl import load_workbook

BASE_DIR = Path(__file__).parent
app = FastAPI(title="GSRTC Proposal Generator")
templates = Jinja2Templates(directory=BASE_DIR / "templates")


def get_location_data():
    """Read pricing from the rate card so the frontend can show prices."""
    wb = load_workbook(BASE_DIR / "assets" / "ratecard.xlsx", data_only=True)
    ws = wb["Sheet1"]

    locations = []
    for loc in LOCATIONS_IN_ORDER:
        row = LOCATION_EXCEL_ROW[loc]
        locations.append({
            "name": loc,
            "screens": int(ws.cell(row=row, column=5).value or 0),
            "unit_inch": str(ws.cell(row=row, column=3).value or ""),
            "grade": str(ws.cell(row=row, column=4).value or ""),
            "total_month": int(ws.cell(row=row, column=12).value or 0),
            "special_offer": int(ws.cell(row=row, column=13).value or 0),
        })
    return locations


@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    locations = get_location_data()
    return templates.TemplateResponse(
        request=request,
        name="index.html",
        context={"locations": locations},
    )


class ProposalRequest(BaseModel):
    company: str = ""
    client_name: str = ""
    client_mobile: str = ""
    client_email: str = ""
    sender_name: str = ""
    sender_mobile: str = ""
    sender_email: str = ""
    locations: List[str]
    months: int = 1
    spot_duration_sec: int = 10
    location_overrides: dict = {}


@app.post("/generate/pptx")
async def gen_pptx(req: ProposalRequest):
    # Require at least company OR client name
    if not req.client_name.strip() and not req.company.strip():
        raise HTTPException(400, "Either company name or contact person is required")
    if not req.locations:
        raise HTTPException(400, "Select at least one location")

    try:
        data = generate_pptx(
            client_name=req.client_name,
            company=req.company,
            mobile=req.client_mobile,
            email=req.client_email,
            selected_locations=req.locations,
            sender_name=req.sender_name,
            sender_mobile=req.sender_mobile,
            sender_email=req.sender_email,
            months=req.months,
            spot_duration_sec=req.spot_duration_sec,
        )
    except Exception as e:
        raise HTTPException(500, f"Failed to generate PPT: {e}")

    base = req.company.strip() or req.client_name.strip() or "Client"
    filename = f"GSRTC_Proposal_{sanitize_filename(base)}.pptx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.presentationml.presentation",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.post("/generate/xlsx")
async def gen_xlsx(req: ProposalRequest):
    if not req.client_name.strip() and not req.company.strip():
        raise HTTPException(400, "Either company name or contact person is required")
    if not req.locations:
        raise HTTPException(400, "Select at least one location")

    try:
        data = generate_xlsx(
            client_name=req.client_name,
            company=req.company,
            mobile=req.client_mobile,
            email=req.client_email,
            selected_locations=req.locations,
            sender_name=req.sender_name,
            sender_mobile=req.sender_mobile,
            sender_email=req.sender_email,
            months=req.months,
            spot_duration_sec=req.spot_duration_sec,
            location_overrides=req.location_overrides,
        )
    except Exception as e:
        raise HTTPException(500, f"Failed to generate Excel: {e}")

    base = req.company.strip() or req.client_name.strip() or "Client"
    filename = f"GSRTC_RateCard_{sanitize_filename(base)}.xlsx"
    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


@app.get("/health")
async def health():
    return {"status": "ok"}


# ---------- Form-based download endpoint (works in mobile Safari) ----------
# Mobile Safari doesn't reliably download files fetched via JS fetch + blob.
# A native form POST that returns the file triggers Safari's normal download flow.

import json as _json


@app.post("/download/{kind}")
async def download_file_form(kind: str,
                             company: str = Form(""),
                             client_name: str = Form(""),
                             client_mobile: str = Form(""),
                             client_email: str = Form(""),
                             sender_name: str = Form(""),
                             sender_mobile: str = Form(""),
                             sender_email: str = Form(""),
                             locations: str = Form("[]"),
                             months: int = Form(1),
                             spot_duration_sec: int = Form(10),
                             location_overrides: str = Form("{}")):
    """
    Same as /generate/{kind} but accepts form-encoded data.
    `locations` and `location_overrides` are JSON-encoded strings.
    """
    if kind not in ("pptx", "xlsx"):
        raise HTTPException(404, "Unknown file kind")

    if not client_name.strip() and not company.strip():
        raise HTTPException(400, "Either company name or contact person is required")

    try:
        locations_list = _json.loads(locations) if locations else []
        overrides_dict = _json.loads(location_overrides) if location_overrides else {}
    except _json.JSONDecodeError:
        raise HTTPException(400, "Invalid locations or overrides format")

    if not locations_list:
        raise HTTPException(400, "Select at least one location")

    try:
        if kind == "pptx":
            data = generate_pptx(
                client_name=client_name, company=company,
                mobile=client_mobile, email=client_email,
                selected_locations=locations_list,
                sender_name=sender_name, sender_mobile=sender_mobile,
                sender_email=sender_email,
                months=months, spot_duration_sec=spot_duration_sec,
            )
            mime = "application/vnd.openxmlformats-officedocument.presentationml.presentation"
            prefix = "GSRTC_Proposal_"
        else:
            data = generate_xlsx(
                client_name=client_name, company=company,
                mobile=client_mobile, email=client_email,
                selected_locations=locations_list,
                sender_name=sender_name, sender_mobile=sender_mobile,
                sender_email=sender_email,
                months=months, spot_duration_sec=spot_duration_sec,
                location_overrides=overrides_dict,
            )
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            prefix = "GSRTC_RateCard_"
    except Exception as e:
        raise HTTPException(500, f"Failed to generate {kind}: {e}")

    base = company.strip() or client_name.strip() or "Client"
    filename = f"{prefix}{sanitize_filename(base)}.{kind}"
    return Response(
        content=data,
        media_type=mime,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
